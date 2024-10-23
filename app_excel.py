import openpyxl


# wb = openpyxl.Workbook()
wb = openpyxl.load_workbook('transactions.xlsx')
print(wb.sheetnames)

sheet = wb['Sheet1']
#index based so can place it where ever
# wb.create_sheet('Sheet2, 0)
# wb.remove_sheet(sheet)

cell = sheet['a1']
column = sheet['a']
cells = sheet['a:c']
all_used_cells = sheet['a1:c4']
rows_of_cells = sheet[1:3]
print(cells)
# print(cell.value)
# print(cell.row)
# print(cell.column)
# print(cell.coordinate)
# cell2 = sheet.cell(row=1, column=1)
# print(cell == cell2)

# print(sheet.max_row)
# print(sheet.max_column)

# range because its a number, 1 because it does not use zero index and +1 because we need the last row
# for row in range(1, sheet.max_row + 1):
#     for column in range(1, sheet.max_column + 1):
#         cell = sheet.cell(row, column)
#         print(cell.value)

# inserts a row at the end
sheet.append([1,2,3])
wb.save('transactions2.xlsx')